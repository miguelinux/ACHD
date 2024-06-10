from flask import Blueprint, render_template, redirect, request, jsonify,Response,flash
from functions import verificate_session
from models.tables_db import Usuarios, Materias, Aulas, Asignaciones, Ciclos, Grupo
from models.tables_db import DocenteCarreras, MateriasCarreras, Disponibilidades, GrupoSemestre
from extensions import db
import json, io
from sqlalchemy import asc
from sqlalchemy.orm import joinedload
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo

jefe_bp = Blueprint('jefe', __name__)

@jefe_bp.route("/jefeCarrera")
def jefe_carrera():
    user = verificate_session()
    if user:
        username = user["username"]
        return render_template("jefeCarrera.html", user=username)
    return redirect("/")


@jefe_bp.route("/jefeCarrera/docentes")
def docentes():
    user = verificate_session()
    if user:
        try:
            userid = request.args["userid"]
        except KeyError:
            userid = None
        username = user["username"]
        carrera = user["carrera"]
        
        docentes = (
            DocenteCarreras.query
            .join(Usuarios, DocenteCarreras.usuario_id == Usuarios.id)
            .filter(DocenteCarreras.carrera_id == carrera)
            .options(joinedload(DocenteCarreras.usuario))
            .order_by(Usuarios.nombre)
            .all()
        )
        return render_template("docentes.html", user=username, docentes=docentes, userid=userid)
    return redirect("/")


@jefe_bp.route("/horarioJ")
def horarioJefe():
    user = verificate_session()
    if user:
        try:
            user_id = request.args["userid"]
            if not user_id or user_id == "None":
                return """
                <div style="color: white; font-weight: bold; font-family: Arial, sans-serif; background-color: red; padding: 10px; border-radius: 5px;">
                    NO HA SELECCIONADO NINGÚN DOCENTE
                </div>
                """
        except KeyError:
            return """
            <div style="color: white; font-weight: bold; font-family: Arial, sans-serif; background-color: red; padding: 10px; border-radius: 5px;">
                NO HA SELECCIONADO NINGÚN DOCENTE
            </div>
            """
        ciclo = Ciclos.query.filter_by(actual=True).first()        
        dispo = Disponibilidades.query.filter_by(usuario_id=user_id, ciclo_id=ciclo.id).first()
        
        try:
            disponibilidad = dispo.horas
        except AttributeError:
            return """
            <div style="color: white; font-weight: bold; font-family: Arial, sans-serif; background-color: red; padding: 10px; border-radius: 5px;">
                EL HORARIO AÚN NO HA SIDO CARGADO POR EL DOCENTE
            </div>
            """
        if not disponibilidad:
            return """
            <div style="color: white; font-weight: bold; font-family: Arial, sans-serif; background-color: red; padding: 10px; border-radius: 5px;">
                EL HORARIO AÚN NO HA SIDO CARGADO POR EL DOCENTE
            </div>
            """

        return render_template("horarioJ.html", user=user, disponibilidad=disponibilidad, usuario=dispo)
    return redirect("/")



@jefe_bp.route("/jefeCarrera/materias")
def materias():
    user = verificate_session()
    if user:
        username = user["username"]
        carrera = user["carrera"]
        
        materias = (
            MateriasCarreras.query
            .join(Materias, MateriasCarreras.materia_id == Materias.id)
            .filter(MateriasCarreras.carrera_id == carrera)
            .options(joinedload(MateriasCarreras.materia))
            .order_by(Materias.semestre)
            .all()
        )
        return render_template("materias.html", user=username, materias=materias)
    return redirect("/")

@jefe_bp.route("/jefeCarrera/asignacion")
def asignacion():
    user = verificate_session()
    if user:
        username = user["username"]
        carrera = user["carrera"]
        ciclo = Ciclos.query.filter_by(actual=True).first()

        docentes = (
            DocenteCarreras.query
            .filter(DocenteCarreras.carrera_id == carrera)
            .outerjoin(Disponibilidades, DocenteCarreras.usuario_id == Disponibilidades.usuario_id)
            .filter(Disponibilidades.usuario_id != None,
                    Disponibilidades.ciclo_id == ciclo.id)
            .all()
        )

        asignaturas = MateriasCarreras.query.filter_by(carrera_id=carrera)
        aulas = Aulas.query.all()

        # Obtener los grupos y sus semestres
        grupos_semestre = GrupoSemestre.query.all()
        grupos_con_semestre = []
        for gs in grupos_semestre:
            grupo = Grupo.query.get(gs.grupo_id)
            grupos_con_semestre.append({
                'id': gs.id,
                'identificador': grupo.identificador,
                'semestre': gs.semestre
            })

        return render_template("asignacion.html", user=username, asignaturas=asignaturas, docentes=docentes, aulas=aulas, grupos=grupos_con_semestre)
    return redirect("/")

@jefe_bp.route("/setAsignacion", methods=["POST"])
def set_asignacion():
    user = verificate_session()
    if user:
        user_carrera = user["carrera"]
        asignacion_propuesta = request.json
        turno = asignacion_propuesta.pop("turno")
        ciclo = Ciclos.query.filter_by(actual=True).first()
        grupsemestre = GrupoSemestre.query.filter_by(id=turno).first()
        asignacion = Asignaciones.query.filter_by(carrera_id=user_carrera, grupo_id=grupsemestre.id, ciclo_id=ciclo.id).first()
        print(turno)
        if not asignacion:
            asignacion = Asignaciones(
                carrera_id=user_carrera,
                grupo_id=turno,
                ciclo_id=ciclo.id,
                horario=json.dumps(asignacion_propuesta.get("asignacion", {}))
            )
            horariov = None
            db.session.add(asignacion)
        else:
            horariov = json.loads(asignacion.horario)
            asignacion.horario = json.dumps(asignacion_propuesta.get("asignacion", {}))

        db.session.commit()
        horarion = json.loads(asignacion.horario)
        semestre = horarion.get("docente", [])
        cell_ids = horarion.get("cell_ids", [])

        semestre_actualizados = 0

        if horariov:
            actualizar = detectar_cambios(horarion, horariov)
            semestreold = actualizar.get("docente", [])
            cell_idsold = actualizar.get("cell_ids", [])

            for i, doc_id in enumerate(semestreold):
                if doc_id != '':
                    cell_id = int(cell_idsold[i])
                    if update_disp(doc_id, [cell_id], 1):
                        semestre_actualizados += 1

        for i, doc_id in enumerate(semestre):
            if doc_id != '':
                cell_id = int(cell_ids[i])
                if update_disp(doc_id, [cell_id], 3):
                    semestre_actualizados += 1

        if semestre_actualizados > 0:
            return jsonify({"success": True, "message": f"Disponibilidad actualizada para {semestre_actualizados} docente(s)"})
        else:
            return jsonify({"success": False, "message": "No se encontraron semestre con disponibilidad para actualizar"})

    return redirect("/")

def detectar_cambios(jsnew, jsold):
    cambios = {"docente": [], "cell_ids": []}
    for old_docente, new_docente, cell_id in zip(jsold["docente"], jsnew["docente"], jsnew["cell_ids"]):
        if old_docente != new_docente and old_docente != "":
            cambios["docente"].append(old_docente)
            cambios["cell_ids"].append(cell_id)
    return cambios

def update_disp(user_id, indices_to_update, value):
    ciclo = Ciclos.query.filter_by(actual=True).first()
    dispo= Disponibilidades.query.filter_by(usuario_id=user_id,ciclo_id=ciclo.id).first()
    
    if dispo:
        current_availability = json.loads(dispo.horas)["disponibilidad"]
        for idx in indices_to_update:
            current_availability[idx] = value

        result_dict = {"disponibilidad": current_availability}
        result_json = json.dumps(result_dict)

        dispo.horas = result_json
        db.session.commit()
        return True
    return False

@jefe_bp.route("/jefeCarrera/semestres")
def semestres():
    user = verificate_session()
    if user:
        username = user["username"]
        carrera = user["carrera"]
        docentes = DocenteCarreras.query.filter_by(carrera_id=carrera).all()
        asignaturas = MateriasCarreras.query.filter_by(carrera_id=carrera).all()
        aula = Aulas.query.all()
        
        grupos_semestre = GrupoSemestre.query.all()
        grupos_con_semestre = []
        for gs in grupos_semestre:
            grupo = Grupo.query.get(gs.grupo_id)
            grupos_con_semestre.append({
                'id': gs.id,
                'identificador': grupo.identificador,
                'semestre': gs.semestre
            })
        
        return render_template("semestres.html", user=username, asignaturas=asignaturas, docentes=docentes, aulas=aula,grupos=grupos_con_semestre)
    return redirect("/")

@jefe_bp.route("/jefeCarrera/grupos")
def grupos():
    user = verificate_session()
    if user:
        username = user["username"]
        carrera = user["carrera"]
        ciclo = Ciclos.query.filter_by(actual=True).first()

        # Realizar join entre Grupo y GrupoSemestre
        resultados = db.session.query(Grupo, GrupoSemestre).outerjoin(GrupoSemestre, Grupo.id == GrupoSemestre.grupo_id)\
            .filter(Grupo.carrera_id == carrera, Grupo.ciclo_id == ciclo.id).all()
        
        # Agrupar los semestres por grupo
        grupos = {}
        for grupo, grupo_semestre in resultados:
            if grupo.id not in grupos:
                grupos[grupo.id] = {
                    'identificador': grupo.identificador,
                    'semestres': []
                }
            if grupo_semestre:
                grupos[grupo.id]['semestres'].append(grupo_semestre.semestre)

        for grupo_id in grupos:
            grupos[grupo_id]['semestres'].sort()

        return render_template("grupos.html", user=username, grupos=grupos)
    return redirect("/")


@jefe_bp.route("/crear_grupo", methods=["POST"])
def crear_grupo():
    """
    Metodo para agregar un grupo a la base de datos
    """
    user = verificate_session()
    if user:
        carrera = user["carrera"]
        ciclo = Ciclos.query.filter_by(actual=True).first()
        form = request.form
        identificador = form.get("identificador")
        
        # Validar si ya existe un grupo con el mismo identificador en la misma carrera y ciclo
        grupo_existente = Grupo.query.filter_by(identificador=identificador, carrera_id=carrera, ciclo_id=ciclo.id).first()
        
        if grupo_existente:
            flash("Ya existe un grupo con este identificador en la carrera y ciclo actuales.", "error")
            return redirect("/jefeCarrera/grupos")

        nuevo_grupo = Grupo(
            identificador=identificador,
            carrera_id=carrera,
            ciclo_id=ciclo.id
        )
        db.session.add(nuevo_grupo)
        db.session.commit()
        flash("Grupo creado exitosamente.", "success")
        return redirect("/jefeCarrera/grupos")
    else:
        return redirect("/")

def grupo_semestre(grupoid, semestre):
    # Verificar si el docente está asociado a la carrera en la base de datos
    return bool(GrupoSemestre.query.filter_by(grupo_id=grupoid, semestre=semestre).first())

@jefe_bp.route("/jefe/asignar_grupo_semestre/<int:grupoId>", methods=['GET'])
def grupos_semestre(grupoId):
    user = verificate_session()
    if user:
        username = user["username"]
        carrera = user["carrera"]
        ciclo=Ciclos.query.filter_by(actual=True).first()
        grupos = Grupo.query.filter_by(carrera_id=carrera,ciclo_id=ciclo.id).all()
        return render_template("grupo_semestre.html", user=username, grupos=grupos,
                    grupoId=grupoId,grupo_semestre=grupo_semestre)
    return redirect("/")


@jefe_bp.route("/guardar_gruposemestre", methods=['POST'])
def guardar_gruposemestre():
    user = verificate_session()
    if user:
        data = request.json
        grupoId = data["grupoId"]
        semestre_seleccionados = data["semestreSeleccionados"]
        semestre_deseleccionados = data["semestreDeseleccionados"]

        if semestre_deseleccionados:
            GrupoSemestre.query.filter(
                GrupoSemestre.grupo_id == grupoId,
                GrupoSemestre.semestre.in_(semestre_deseleccionados)
            ).delete(synchronize_session=False)

        if semestre_seleccionados:
            for semestre in semestre_seleccionados:
                if not GrupoSemestre.query.filter_by(grupo_id=grupoId, semestre=semestre).first():
                    nueva_relacion = GrupoSemestre(grupo_id=grupoId, semestre=semestre)
                    db.session.add(nueva_relacion)
        db.session.commit()

        return jsonify({"success": True})
    return redirect("/")


@jefe_bp.route("/delete/grupo", methods=["POST"])
def delete_grupo():
    """
    Método para borrar un grupo de la base de datos
    """
    user = verificate_session()
    if user:
        grupoid = request.form.get("id")
        grupo = Grupo.query.get(grupoid)
        if grupo:
            db.session.delete(grupo)
            db.session.commit()
        return redirect("/jefeCarrera/grupos")
    return redirect("/")



@jefe_bp.route("/export/grupos", methods=["POST"])
def exportar():
    """
    Método para exportar
    """
    user = verificate_session()
    if user:
        carrera = user["carrera"]
        ciclo = Ciclos.query.filter_by(actual=True).first()
        asignaciones = Asignaciones.query.filter_by(carrera_id=carrera, ciclo_id=ciclo.id).all()
        aulas = Aulas.query.all()
        aula_dict = {aula.id: aula.nombre for aula in aulas}
        docentes_carrera = DocenteCarreras.query.filter_by(carrera_id=carrera).all()
        usuario_ids = [docente.usuario_id for docente in docentes_carrera]
        usuarios = Usuarios.query.filter(Usuarios.id.in_(usuario_ids)).all()
        docente_dict = {usuario.id: f"{usuario.nombre} {usuario.apellido_pat}\n" for usuario in usuarios}
        materias_carrera = MateriasCarreras.query.filter_by(carrera_id=carrera).all()
        materia_ids = [materia.materia_id for materia in materias_carrera]
        materias = Materias.query.filter(Materias.id.in_(materia_ids)).all()
        materia_dict = {materia.id: f"{materia.nombre}\n" for materia in materias}
        grupos_semestre = GrupoSemestre.query.all()
        grupo_dict = {
            grupo_semestre.id: {
                "grupo": Grupo.query.filter_by(id=grupo_semestre.grupo_id).first().identificador,
                "semestre": grupo_semestre.semestre
            }
            for grupo_semestre in grupos_semestre
        }
        grupos_asignaciones = {}
        for asignacion in asignaciones:
            grupo_id = asignacion.grupo_id
            if grupo_id not in grupos_asignaciones:
                grupos_asignaciones[grupo_id] = []
            grupos_asignaciones[grupo_id].append(asignacion)
        
        # Crear un libro de trabajo de Excel
        wb = Workbook()
        ws_docentes = wb.active
        ws_docentes.title = "Docentes"

        # Inicializar el conteo de horas por docente
        horas_docentes = {docente_id: 0 for docente_id in usuario_ids}
        
        semestres_dict = {i: [] for i in range(1, 10)}  # Crear un diccionario para los semestres del 1 al 9
        
        for grupo_id, asignaciones_grupo in grupos_asignaciones.items():
            grupo_info = grupo_dict.get(grupo_id, {"grupo": "Desconocido", "semestre": "Desconocido"})
            semestre = grupo_info["semestre"]
            semestres_dict[semestre].append((grupo_id, asignaciones_grupo))
        
        for semestre, grupos in semestres_dict.items():
            if grupos:
                # Crear una hoja para cada semestre
                ws = wb.create_sheet(title=f"Semestre {semestre}")
                
                current_row = 1  # Mantener la fila actual para separar las tablas

                for grupo_id, asignaciones_grupo in grupos:
                    grupo_info = grupo_dict.get(grupo_id, {"grupo": "Desconocido", "semestre": "Desconocido"})
                    grupo_nombre = grupo_info["grupo"]
                    
                    # Escribir la información del grupo y semestre en la hoja
                    ws.cell(row=current_row, column=1, value=f"Grupo {grupo_nombre} - Semestre {semestre}")
                    current_row += 1
                    dias_semana = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado"]
                    for col, dia in enumerate(dias_semana, start=1):
                        ws.cell(row=current_row, column=col, value=dia)
                    current_row += 1  # Mover a la fila siguiente
                    
                    # Inicializar una lista para contener los valores de la tabla
                    table_values = [["" for _ in range(6)] for _ in range(8)]  # 6 columnas, 8 filas
                    
                    for asignacion in asignaciones_grupo:
                        horario_json = asignacion.horario
                        horario_dict = json.loads(horario_json)
                        docentes = horario_dict["docente"]
                        materias_ids = horario_dict["materia"]
                        aulas_ids = horario_dict["aula"]
                        num_items = min(len(docentes), len(materias_ids), len(aulas_ids))
                        for i in range(num_items):
                            docente_id = docentes[i].strip()
                            materia_id = materias_ids[i].strip()
                            aula_id = aulas_ids[i].strip()
                            aula_nombre = aula_dict.get(int(aula_id), "") if aula_id.isdigit() else ""
                            docente_nombre = docente_dict.get(int(docente_id), "") if docente_id.isdigit() else ""
                            materia_nombre = materia_dict.get(int(materia_id), "") if materia_id.isdigit() else ""
                            cell_value = f"{docente_nombre}{materia_nombre}{aula_nombre}"
                            row_index = i // 6
                            col_index = i % 6
                            table_values[row_index][col_index] = cell_value
                            
                            # Incrementar el conteo de horas del docente
                            if docente_id.isdigit():
                                horas_docentes[int(docente_id)] += 1
                    
                    # Escribir los valores de la tabla en la hoja
                    for row_values in table_values:
                        for col, value in enumerate(row_values, start=1):
                            ws.cell(row=current_row, column=col, value=value)
                        current_row += 1
                    
                    # Crear la tabla en Excel
                    table_end_row = current_row - 1  # La fila final de la tabla
                    tab = Table(displayName=f"TablaDatos{grupo_id}", ref=f"A{current_row-9}:F{table_end_row}")
                    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
                    tab.tableStyleInfo = style
                    ws.add_table(tab)

                    # Aplicar el formato de alineación para permitir saltos de línea en las celdas
                    for row in ws.iter_rows(min_row=current_row-9, max_col=6, max_row=table_end_row):
                        for cell in row:
                            cell.alignment = Alignment(wrap_text=True)
                    
                    # Añadir una fila en blanco entre tablas de grupos diferentes
                    current_row += 1
                
        # Escribir los valores en la hoja "Docentes"
        ws_docentes.append(["Docente", "Horas"])
        for docente_id, horas in horas_docentes.items():
            ws_docentes.append([docente_dict.get(docente_id, "Desconocido"), horas])
        
        # Guardar el libro de trabajo en un objeto de archivo en memoria
        excel_content = io.BytesIO()
        wb.save(excel_content)
        excel_content.seek(0)

        # Configurar la respuesta HTTP para devolver el archivo Excel
        response = Response(excel_content.getvalue(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response.headers['Content-Disposition'] = 'attachment; filename=exportacion_grupos.xlsx'
        return response

    return redirect("/")

@jefe_bp.route("/export/disponibilidades", methods=["POST"])
def exportar_disponibilidad():
    """
    Método para exportar disponibilidades de los docentes de una carrera en el ciclo actual.
    """
    user = verificate_session()
    if user:
        carrera_id = user["carrera"]
        ciclo_actual = Ciclos.query.filter_by(actual=True).first()
        
        if not ciclo_actual:
            return jsonify({"error": "No se encontró el ciclo actual"}), 404
        
        # Obtener los usuarios vinculados a la carrera y ordenarlos por apellido
        docentes_carrera = DocenteCarreras.query.filter_by(carrera_id=carrera_id).all()
        usuario_ids = [docente.usuario_id for docente in docentes_carrera]
        
        # Filtrar usuarios que tienen disponibilidades en el ciclo actual y ordenarlos por apellido
        disponibilidades = (Disponibilidades.query
                            .join(Usuarios)
                            .filter(
                                Disponibilidades.usuario_id.in_(usuario_ids),
                                Disponibilidades.ciclo_id == ciclo_actual.id
                            )
                            .order_by(asc(Usuarios.apellido_pat), asc(Usuarios.apellido_mat), asc(Usuarios.nombre))
                            .options(joinedload(Disponibilidades.usuario))
                            .all())
        
        # Crear un archivo Excel con openpyxl
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Disponibilidades"
        
        # Escribir el encabezado
        sheet.append(["Docente", "Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado"])
        
        for disponibilidad in disponibilidades:
            # Interpretar el JSON de disponibilidad
            disponibilidades_json = json.loads(disponibilidad.horas)
            disponibilidades_lista = disponibilidades_json["disponibilidad"]
            
            # Nombre completo del usuario
            nombre_completo = f"{disponibilidad.usuario.apellido_pat} {disponibilidad.usuario.apellido_mat} {disponibilidad.usuario.nombre}"
            
            # Calcular los rangos de disponibilidad por día
            rangos_disponibilidad = [""] * 7
            for dia in range(7):
                rangos_dia = []
                inicio_disponibilidad = None
                fin_disponibilidad = None
                for hora in range(15):
                    indice = 15 * dia + hora
                    if indice < len(disponibilidades_lista) and disponibilidades_lista[indice] in [1, 3]:
                        if inicio_disponibilidad is None:
                            inicio_disponibilidad = hora
                        fin_disponibilidad = hora + 1
                    elif inicio_disponibilidad is not None:
                        rangos_dia.append(f"{7 + inicio_disponibilidad}:00-{7 + fin_disponibilidad}:00")
                        inicio_disponibilidad = None
                if inicio_disponibilidad is not None:
                    rangos_dia.append(f"{7 + inicio_disponibilidad}:00-{7 + fin_disponibilidad}:00")
                rangos_disponibilidad[dia] = ", ".join(rangos_dia)
            
            # Escribir los datos en el archivo Excel
            sheet.append([nombre_completo] + rangos_disponibilidad)
        
        # Guardar el archivo Excel en memoria
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        # Crear la respuesta usando Response
        response = Response(output.getvalue(),
                            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                            headers={"Content-Disposition": "attachment;filename=disponibilidades.xlsx"})
        
        return response

    return redirect("/")